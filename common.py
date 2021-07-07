
# %%
# NOTE: openpyxl throws an error when loading an xlsx with MS Sans Serif in the styles.xml.  Need to delete that entry

from abc import ABC, abstractmethod
from copy import copy

import pandas as pd
import openpyxl as xl
from openpyxl.styles import Side
from pandas.core.frame import DataFrame


# %%

class TableGenerator(ABC):

    subjgrade_labels = {'M': 'mathematics', 'R': 'reading'}

    def __init__(self, infile: str) -> None:
        self._data = pd.read_excel(infile)

        # Optional data post-processing
        self._postproc_data(self._data)

        self._data = self._data[self._data['IN_SNAKECHART_FILE'] == 'YES']
        self._cons_set = {c for c in self._data['Consortia'] if not pd.isna(c)}

        self.state_data = self._data.loc[self._data['state'].isin(
            self._cons_set) == False]
        self.cons_data = self._data.loc[self._data['state'].isin(
            self._cons_set)]

        self.states = self.state_data.state.unique()
        self.consortia = self.cons_data.state.unique()
        self.years = self._data.year.unique()

        for arr in self.states, self.consortia, self.years:
            arr.sort()

    def _postproc_data(self, input_data: DataFrame):
        pass

    @abstractmethod
    def map_row(self, *args, **kwargs):
        raise NotImplementedError

    @abstractmethod
    def generate(self, *args, **kwargs):
        raise NotImplementedError


class StateTableGenerator(TableGenerator):

    col_offsets = {'R': 3, 'M': 7}

    def map_row(self, nse, nse_se, nse_re, mark, *args, **kwargs):
        nse_nan = pd.isna(nse)
        return tuple((
            '\u2013' if nse_nan else nse,
            '†' if nse_nan else (
                '\u2013' if pd.isna(nse_se) else nse_se),
            '†' if nse_nan else (
                '\u2013' if pd.isna(nse_re) else nse_re),
            '!' if mark == '!' else ''
        ))

    def generate(self, state: str) -> xl.Workbook:

        wb = xl.load_workbook('source_state.xlsx')
        df = self.state_data[self.state_data['state'] == state]

        # Replace state abbr in sheet names
        for sheet in wb:
            sheet.title = sheet.title.replace('_ST_')

        for grade in '4', '8':
            ws = wb['G'+grade]
            ws['B5'].value = ws['B5'].value.replace('[STATE_NAME]', state)

            data = df.loc[df['subjgrade'].str.endswith(
                grade)].sort_values(['year', 'subjgrade'])

            # Populate years (needs error checking?)
            for i, year in enumerate(self.years):
                ws.cell(i+9, 2).value = year

            for subj in 'R', 'M':

                j = self.col_offsets[subj]

                rows = data[data['subjgrade'] == subj +
                            grade].apply(lambda r: self.map_row(**r), axis=1)

                for i, row_data in enumerate(rows):
                    for col in range(len(row_data)):
                        cell = ws.cell(i+9, col+j)
                        cell.value = row_data[col]

            # Hide unneeded notes
            ws.row_dimensions[16].hidden = not pd.isna(
                data[['nse_se', 'nse_re']]).values.any()
            ws.row_dimensions[17].hidden = not pd.isna(
                data['nse']).values.any()
            ws.row_dimensions[18].hidden = not (
                data['mark'] == '!').values.any()

        return wb


class SnakeTableGenerator(TableGenerator):

    letters = {'R4': 'a', 'M4': 'b', 'R8': 'c', 'M8': 'd'}

    def _postproc_data(self, input_data: DataFrame):
        # NOTE: Need to update IN_SNAKECHART_FILE variable to ensure blank rows for
        #       2015 and 2019 for all states except PR and blank rows for PR Reading
        #       in 2017 and beyond
        input_data.loc[input_data['year'] ==
                       2015, 'IN_SNAKECHART_FILE'] = 'YES'
        input_data.loc[input_data['year'] ==
                       2019, 'IN_SNAKECHART_FILE'] = 'YES'
        input_data.loc[input_data['state'] ==
                       'Puerto Rico', 'IN_SNAKECHART_FILE'] = 'YES'
        input_data.loc[(input_data['state'] == 'Puerto Rico') & (
            input_data['year'] < 2017), 'IN_SNAKECHART_FILE'] = 'NO'

    def map_row(self, state, Consortia, nse, nse_se, nse_re, *args, **kwargs):
        nse_nan = pd.isna(nse)
        cons = '†' if pd.isna(Consortia) else Consortia
        return tuple((
            state,
            cons,
            '\u2013' if nse_nan else nse,
            '†' if nse_nan else (
                '\u2013' if pd.isna(nse_se) else nse_se),
            '†' if nse_nan else (
                '\u2013' if pd.isna(nse_re) else nse_re)
        ))

    def generate(self, year: int, subjgrade: str) -> xl.Workbook:
        wb = xl.load_workbook('source_snakechart.xlsx')
        ws = wb['Sheet1']

        styles = [ws.cell(8, i)._style for i in range(2, 7)]

        subj = self.subjgrade_labels[subjgrade[0]]
        grade = subjgrade[1]
        letter = self.letters[subjgrade]
        title = ws['B5'].value.replace(
            '[YEAR]', str(year)
        ).replace(
            '[LETTER]', letter
        ).replace(
            '[GRADE]', grade
        ).replace(
            '[SUBJECT]', subj
        )
        ws['B5'].value = title

        def apply_border(row_idx):
            for i in range(2, 7):
                new_bd = copy(ws.cell(row_idx, i).border)
                new_bd.bottom = Side('thin', '000000')
                ws.cell(row_idx, i).border = new_bd

        # Extract row data
        data_rows = list()
        end_rowidx = list()
        for df in self.state_data, self.cons_data:

            dta = df[(df['year'] == year) & (
                df['subjgrade'] == subjgrade)].sort_values('state')
            if not len(dta):
                continue

            rows = dta.apply(lambda r: self.map_row(**r), axis=1)
            data_rows.extend(rows)
            end_rowidx.append(len(data_rows)+7)

        # Insert rows into sheet
        for i, row in enumerate(data_rows):
            for col in range(len(row)):
                cell = ws.cell(i+8, col+2)
                cell.value = row[col]
                cell._style = copy(styles[col])

        # Format bottom borders
        for idx in end_rowidx:
            apply_border(idx)

        return wb
