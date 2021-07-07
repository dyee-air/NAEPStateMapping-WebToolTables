# %%
from copy import copy
from typing import Any, Collection, Sequence
import pandas as pd
import openpyxl as xl
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

DATAFILE = "StateMappingResults.csv"
OUTPATH = 'output/state'

# %%


class TableData:

    def __init__(self, data_path: str) -> None:
        self._data = pd.read_csv(data_path)

    def get_state_data(self, state: str, grade: int = None):
        state_rows = self._data[(self._data.state == state) & (
            self._data.year >= 2007) & (self._data.is_consortium == False)]

        if grade:
            return state_rows[state_rows.subjgrade.str.endswith(f'{grade}')]

        return state_rows

    def get_state_rows(self, state: str, grade: int):
        dta = self.get_state_data(
            state, grade).sort_values(['year', 'subjgrade'])

        outdf = self._reshape_rows(dta)
        return list(dataframe_to_rows(outdf, header=False, index=False))

    def _reshape_rows(self, df):
        dta_r = df[df.subjgrade.str.startswith(
            'R')][['year', 'nse', 'nse_se', 'nse_re']]
        dta_m = df[df.subjgrade.str.startswith(
            'M')][['year', 'nse', 'nse_se', 'nse_re']]
        for sub_dta in dta_r, dta_m:
            self._sanitize_rows(sub_dta)
        return pd.merge(dta_r, dta_m, on='year', suffixes=['_R', '_M'])

    def _sanitize_rows(self, df):
        df['mark'] = df.apply(
            lambda r: '!' if r.nse_re >= 0.5 else '', axis=1)
        df.loc[pd.isna(df.nse), 'nse'] = '\u2013'
        df.loc[(df.nse == '\u2013') | (pd.isna(df.nse_se)), 'nse_se'] = '†'
        df.loc[(df.nse == '\u2013') | (pd.isna(df.nse_re)), 'nse_re'] = '†'

    def get_consortium_rows(self, state: str, grade: int):
        dta = self.get_state_data(
            state, grade).sort_values(['year', 'subjgrade'])

        # Need to handle years when state has consortium in one
        # subject but not other
        cons_yrs = dta.loc[pd.notna(dta.consortium), 'year'].unique()
        query_dta = dta[dta.year.isin(
            cons_yrs)][['year', 'subjgrade', 'consortium']]

        if query_dta.empty:
            return []

        dta = pd.merge(query_dta, self._data,
                       left_on=['year', 'subjgrade', 'consortium'],
                       right_on=['year', 'subjgrade', 'state'],
                       how='left')
        outdf = self._reshape_rows(
            dta[['year', 'subjgrade', 'nse', 'nse_se', 'nse_re']])
        outdf = pd.merge(query_dta.loc[pd.notna(query_dta.consortium), [
                         'year', 'consortium']].drop_duplicates(), outdf, on='year')
        return list(dataframe_to_rows(outdf, header=False, index=False))

# %%


# %%

CONS_NOTES = {
    'ACT': 'ACT refers to ACT Aspire.',
    'NECAP': 'NECAP refers to New England Common Assessment Program.',
    'PARCC': 'PARCC refers to Partnership for Assessment of Readiness for College and Careers.',
    'SBAC': 'SBAC refers to Smarter Balanced Assessment Consortium.'
}

NO_BORDER = Border(left=Side(border_style=None),
                   right=Side(border_style=None),
                   top=Side(border_style=None),
                   bottom=Side(border_style=None))


def get_consortium_text(consortia: Collection[str]):
    return ' '.join(CONS_NOTES.get(consortium, '') for consortium in consortia)


def remove_rows(ws, start, num=1):
    ws.delete_rows(start, num)
    for i in range(start, start+num+1):
        merged = [
            rng for rng in ws.merged_cells.ranges if rng.coord.startswith(f'A{i}')]
        for rng in merged:
            ws.merged_cells.remove(rng.coord)

        ws.row_dimensions[i].height = None


def get_note_row(ws, rownum: int):
    row = list(ws.rows)[rownum-1]
    cell = row[0]
    return {'value': cell.value,
            'height': ws.row_dimensions[cell.row].height,
            'font': copy(cell.font),
            'alignment': copy(cell.alignment)}


def write_notes(ws, notes, start_row=14):
    end_col = 'I' if ws.max_column == 9 else 'J'
    for i, note in enumerate(notes):
        rownum = start_row+i

        tgt_cell = ws[f'A{rownum}']
        tgt_cell.value = note['value']
        tgt_cell.font = note['font']
        tgt_cell.alignment = note['alignment']
        tgt_cell.border = copy(NO_BORDER)
        ws.row_dimensions[rownum].height = note['height']
        ws.merge_cells(f'A{rownum}:{end_col}{rownum}')


# STATE TABLE


def year_title_text(years: Sequence[int]) -> str:
    if not years:
        return ''

    yr_str = [str(year) for year in years]

    # Single
    if len(yr_str) == 1:
        return yr_str[0]

    # Two years
    if len(yr_str) == 2:
        return f'{yr_str[0]} and {yr_str[1]}'

    # Three or more years
    return f'Various years, {yr_str[0]}\u2013{yr_str[-1][-2:]}'


def year_note_text(years: Sequence[int], edfacts: bool = False):
    if edfacts:
        yr_str = [f'{year-1}-{str(year)[-2:]}' for year in years]
    else:
        yr_str = [str(year) for year in years]

    if len(yr_str) == 1:
        return yr_str[0]

    if len(yr_str) == 2:
        return ' and '.join(yr_str)

    return ', '.join(yr_str[:-1]) + ', and ' + yr_str[-1]


def edit_notes_PR(note_rows: Sequence):
    txt = note_rows[-1]['value']
    txt = txt.replace(
        '2007, 2009, 2011, 2013, 2015, 2017, and 2019 Reading and ', '2017 and 2019 '
    ).replace(
        '2006–07, 2008–09, 2010–11, 2012–13, 2014–15, 2016–17, ', '2016-17 '
    )
    note_rows[-1]['value'] = txt

    note_rows[-1]['height'] -= 13


def write_table(ws, state: str, state_abbr: str, row_data: Sequence[Sequence[Any]]):

    # Set sheet name
    ws.title = ws.title.replace('_ST_', state_abbr)

    # Set table title
    title_text = ws['A4'].value.replace('[STATE]', state)

    years = [row[0] for row in row_data]
    if '[YEARS]' in title_text:
        title_text = title_text.replace('[YEARS]', year_title_text(years))

    ws['A4'].value = title_text

    # Write data
    for i, rowdta in enumerate(row_data):
        row = list(ws.rows)[i+6]
        for j, dta in enumerate(rowdta):
            row[j].value = dta

    # Get notes
    note_rows = [get_note_row(ws, i+14) for i in range(5)]
    if state_abbr == 'PR':
        edit_notes_PR(note_rows)

    # Add consortium note text
    if '[CONSORTIUM_NOTES]' in note_rows[3]['value'] and (consortia := {r[1] for r in row_data}):
        note_rows[3]['value'] = note_rows[3]['value'].replace(
            '[CONSORTIUM_NOTES]',
            get_consortium_text(consortia)
        )
        if len(consortia) > 1 or 'PARCC' in note_rows[3]['value']:
            note_rows[3]['height'] += 13

    # Add year text for SOURCE note
    if '[YEARS_NAEP]' in note_rows[4]['value']:
        note_rows[4]['value'] = note_rows[4]['value'].replace(
            '[YEARS_NAEP]', year_note_text(years)
        ).replace(
            '[YEARS_EDFACTS]', year_note_text(years, edfacts=True)
        )

    # Flags for inclusion
    flags = [any('\u2013' in row for row in row_data),
             any('†' in row for row in row_data),
             any('!' in row for row in row_data),
             True,
             True]
    note_rows = [note for flag, note in zip(flags, note_rows) if flag]

    # Shrink table
    tbl = list(ws.tables.values())[0]
    tbl.ref = tbl.ref.replace('13', f'{6 + len(row_data)}')
    write_notes(ws, note_rows, start_row=7+len(row_data))

    len_data = 7+len(row_data)+len(note_rows)
    remove_rows(ws, len_data, ws.max_row-len_data)


# %%
td = TableData(DATAFILE)

STATES = td._data.loc[td._data.is_consortium ==
                      False, ['state', 'state_abbr']].drop_duplicates()
# %%
for _, state_row in STATES.iterrows():
    stname = state_row['state']
    stabbr = state_row['state_abbr']
    print(f'State: {stname}')
    WB = xl.load_workbook('template_state.xlsx')
    SHEETS = list(WB)
    for i, sheet in enumerate(SHEETS):
        gr = 8 if i % 2 else 4
        rowfn = td.get_state_rows if i < 2 else td.get_consortium_rows
        rows = rowfn(stname, gr)
        if not rows:
            del WB[sheet.title]
        else:
            write_table(sheet, stname, stabbr, rows)

    WB.save(f'output/state/{stname}.xlsx')

# %%
