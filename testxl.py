# %%
from typing import Optional, Union
import openpyxl as xl
import pandas as pd
import numpy as np

# %%

DATAPATH = 'U:/ESSIN Task 14/Mapping Report/2019/Standard Method/02_Mapping Tool Design/Data Source/SourceTable_States/SAS/Updated data/Long Files/Updated 0208'
OUTPATH = 'U:/ESSIN Task 14/Mapping Report/2019/Standard Method/06_Analysis and Results/Darrick/Web Tables'
dta = dict()
for gr in 'G4', 'G8':
    dta[gr] = pd.read_excel(f'{DATAPATH}/{gr}_long.xlsx')
    for col in 'nse', 'nse_se', 'nse_re':
        dta[gr][col] = pd.to_numeric(dta[gr][col], errors='coerce')

df = dta['G4'].append(dta['G8'])
del dta
# %%


def display_row(year: int,
                nse: float = np.nan,
                nse_se: float = np.nan,
                nse_re: float = np.nan):
    nse_nan = np.isnan(nse)
    return tuple((
        year,
        '-' if nse_nan else nse,
        '†' if nse_nan else (
            '-' if np.isnan(nse_se) else nse_se),
        '†' if nse_nan else (
            '-' if np.isnan(nse_re) else nse_re),
        '!' if nse_re > 0.5 else ''
    ))
# %%


# %%


COL_OFFSETS = {
    'R': 3,
    'M': 7
}

WB = xl.load_workbook('source.xlsx')

sheets = {sg: WB[sg] for sg in ('G4', 'G8')}

STATES = tuple(dict.fromkeys(df['state']))


#%%

for state in STATES:

    for grade in '4', '8':
        ws = sheets['G'+grade]
        ws['B5'].value = ws['B5'].value.replace('[STATE_NAME]', state)

        data = df[(df['subjgrade'].str.endswith(grade)) & (df['state'] == state)]

        for i, year in enumerate(dict.fromkeys(data['year'])):
            ws.cell(i+9, 2).value = year

        for subj in 'R', 'M':

            j = COL_OFFSETS[subj]

            for i, row in enumerate(data[data['subjgrade']==subj+grade].loc[:, 'year':'nse_re'].itertuples()):
                outrow = display_row(*row[1:])
                for col in range(len(outrow[1:])):
                    cell = ws.cell(i+9, col+j)
                    cell.value = outrow[col+1]


        WB.save(f'{OUTPATH}/{state}.xlsx')

# %%
